from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
from pathlib import Path

from openpyxl import load_workbook


SHEET_SPECS = {
    "works": [
        "id",
        "title_zh",
        "title_en",
        "year",
        "material_zh",
        "material_en",
        "project_zh",
        "project_en",
        "size_zh",
        "size_en",
        "location_zh",
        "location_en",
        "description_zh",
        "description_en",
        "inspiration_zh",
        "inspiration_en",
        "cover_image",
        "gallery_images",
    ],
    "exhibitions": [
        "id",
        "title_zh",
        "title_en",
        "year",
        "location_zh",
        "location_en",
        "description_zh",
        "description_en",
        "cover_image",
        "gallery_images",
    ],
    "news": [
        "id",
        "title_zh",
        "title_en",
        "date",
        "content_zh",
        "content_en",
        "related_report_label_zh",
        "related_report_label_en",
        "related_report_url",
        "full_video_label_zh",
        "full_video_label_en",
        "full_video_url",
        "cover_image",
        "gallery_images",
    ],
}

REQUIRED_FIELDS = {
    "works": ["id", "title_zh", "title_en"],
    "exhibitions": ["id", "title_zh", "title_en"],
    "news": ["id", "title_zh", "title_en"],
}

BILINGUAL_FIELD_PAIRS = {
    "works": [
        ("material_zh", "material_en"),
        ("project_zh", "project_en"),
        ("size_zh", "size_en"),
        ("location_zh", "location_en"),
        ("description_zh", "description_en"),
        ("inspiration_zh", "inspiration_en"),
    ],
    "exhibitions": [
        ("location_zh", "location_en"),
        ("description_zh", "description_en"),
    ],
    "news": [
        ("content_zh", "content_en"),
        ("related_report_label_zh", "related_report_label_en"),
        ("full_video_label_zh", "full_video_label_en"),
    ],
}

LINK_FIELD_GROUPS = {
    "news": [
        ("related_report_label_zh", "related_report_label_en", "related_report_url"),
        ("full_video_label_zh", "full_video_label_en", "full_video_url"),
    ]
}

IMAGE_FIELDS = ("cover_image", "gallery_images")


def split_multi_value(value: str | None) -> list[str]:
    if not value:
        return []
    parts = re.split(r"[;\n,]+", value)
    return [part.strip() for part in parts if part.strip()]


def normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def project_root_for_workbook(workbook_path: Path) -> Path:
    workbook_path = workbook_path.resolve()
    if workbook_path.parent.name.lower() == "excel" and len(workbook_path.parents) >= 2:
        return workbook_path.parent.parent
    return workbook_path.parent


def normalize_year_value(value: str, sheet_name: str, row_number: int) -> str:
    if not value:
        return ""
    match = re.search(r"(?<!\d)(\d{4})(?!\d)", value)
    if not match:
        raise ValueError(f"{sheet_name} row {row_number}: invalid year '{value}'")
    return match.group(1)


def normalize_date_value(value: str, sheet_name: str, row_number: int) -> str:
    if not value:
        return ""

    digits = re.findall(r"\d+", value)
    if not digits or len(digits[0]) != 4:
        raise ValueError(f"{sheet_name} row {row_number}: invalid date '{value}'")

    year = int(digits[0])
    if len(digits) == 1:
        return f"{year:04d}"
    if len(digits) == 2:
        month = int(digits[1])
        dt.date(year, month, 1)
        return f"{year:04d}-{month:02d}"

    month = int(digits[1])
    day = int(digits[2])
    dt.date(year, month, day)
    return f"{year:04d}-{month:02d}-{day:02d}"


def normalize_record(record: dict, sheet_name: str, row_number: int) -> dict:
    if sheet_name in {"works", "exhibitions"}:
        record["year"] = normalize_year_value(record.get("year", ""), sheet_name, row_number)
    if sheet_name == "news":
        record["date"] = normalize_date_value(record.get("date", ""), sheet_name, row_number)
    return record


def validate_required_fields(record: dict, sheet_name: str, row_number: int) -> None:
    for field in REQUIRED_FIELDS.get(sheet_name, []):
        if not record.get(field):
            raise ValueError(f"{sheet_name} row {row_number}: missing required field '{field}'")


def validate_bilingual_fields(record: dict, sheet_name: str, row_number: int) -> None:
    for zh_key, en_key in BILINGUAL_FIELD_PAIRS.get(sheet_name, []):
        zh_value = record.get(zh_key, "")
        en_value = record.get(en_key, "")
        if bool(zh_value) != bool(en_value):
            print(f"Warning: {sheet_name} row {row_number}: fields '{zh_key}' and '{en_key}' should be filled together")


def validate_link_groups(record: dict, sheet_name: str, row_number: int) -> None:
    for zh_key, en_key, url_key in LINK_FIELD_GROUPS.get(sheet_name, []):
        values = [record.get(zh_key, ""), record.get(en_key, ""), record.get(url_key, "")]
        if any(values) and not all(values):
            raise ValueError(
                f"{sheet_name} row {row_number}: fields '{zh_key}', '{en_key}', and '{url_key}' must be filled together"
            )


def validate_asset_path(asset_path: str, project_root: Path, sheet_name: str, row_number: int) -> None:
    if not asset_path:
        return
    if re.match(r"^https?://", asset_path, re.IGNORECASE):
        return

    relative_path = Path(asset_path.lstrip("/"))
    candidate = (project_root / relative_path).resolve()
    try:
        candidate.relative_to(project_root.resolve())
    except ValueError as exc:
        raise ValueError(f"{sheet_name} row {row_number}: invalid asset path '{asset_path}'") from exc

    if not candidate.exists():
        raise ValueError(f"{sheet_name} row {row_number}: Missing asset '{asset_path}'")


def validate_assets(record: dict, sheet_name: str, row_number: int, project_root: Path) -> None:
    validate_asset_path(record.get("cover_image", ""), project_root, sheet_name, row_number)
    for path in record.get("gallery_images", []):
        validate_asset_path(path, project_root, sheet_name, row_number)


def validate_unique_ids(records: list[dict], sheet_name: str) -> None:
    seen: dict[str, int] = {}
    for record in records:
        record_id = record.get("id", "")
        row_number = record.get("_row_number", -1)
        if record_id in seen:
            raise ValueError(
                f"{sheet_name} row {row_number}: Duplicate id '{record_id}' (first seen at row {seen[record_id]})"
            )
        seen[record_id] = row_number


def sheet_to_records(workbook_path: Path, sheet_name: str, project_root: Path | None = None) -> list[dict]:
    project_root = project_root or project_root_for_workbook(workbook_path)
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Missing required sheet: {sheet_name}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_cell(value) for value in rows[0]]
    expected = SHEET_SPECS[sheet_name]
    if headers[: len(expected)] != expected:
        raise ValueError(f"Sheet '{sheet_name}' headers do not match expected fields.")

    records = []
    for row_number, raw_row in enumerate(rows[1:], start=2):
        if not raw_row or not any(cell is not None and str(cell).strip() for cell in raw_row):
            continue
        record = {}
        for index, key in enumerate(expected):
            record[key] = normalize_cell(raw_row[index] if index < len(raw_row) else "")
        record["gallery_images"] = split_multi_value(record.get("gallery_images"))
        record["_row_number"] = row_number
        record = normalize_record(record, sheet_name, row_number)
        validate_required_fields(record, sheet_name, row_number)
        validate_bilingual_fields(record, sheet_name, row_number)
        validate_link_groups(record, sheet_name, row_number)
        validate_assets(record, sheet_name, row_number, project_root)
        records.append(record)

    validate_unique_ids(records, sheet_name)
    for record in records:
        record.pop("_row_number", None)
    return records


def export_workbook(workbook_path: Path, output_dir: Path) -> None:
    os.makedirs(output_dir, exist_ok=True)
    project_root = project_root_for_workbook(workbook_path)
    for sheet_name in SHEET_SPECS:
        records = sheet_to_records(workbook_path, sheet_name, project_root=project_root)
        (output_dir / f"{sheet_name}.json").write_text(
            json.dumps(records, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert CMS Excel sheets into JSON data files.")
    parser.add_argument("--workbook", type=Path, required=True, help="Path to excel/content.xlsx")
    parser.add_argument("--output-dir", type=Path, required=True, help="Output directory for JSON files.")
    args = parser.parse_args()

    export_workbook(args.workbook.resolve(), args.output_dir.resolve())
    print(f"Exported JSON to: {args.output_dir.resolve()}")


if __name__ == "__main__":
    main()
