from __future__ import annotations

from pathlib import Path
import re

from openpyxl import Workbook, load_workbook


PACK_ROOT = Path(r"C:\Users\weixi\Desktop\张满堂资料总库_分类版")
SOURCE_WORKBOOK = Path(r"C:\Users\weixi\Desktop\zhangmantang-website\excel\content.xlsx")


def load_records(sheet_name: str, year_key: str, content_key: str):
    wb = load_workbook(SOURCE_WORKBOOK, data_only=True)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item_id = row[idx["id"]]
        if not item_id:
            continue
        gallery_raw = str(row[idx["gallery_images"]] or "").strip()
        gallery = [part.strip() for part in gallery_raw.split(";") if part.strip()]
        rows.append(
            {
                "source_id": str(item_id),
                "year": str(row[idx[year_key]] or "").strip(),
                "title": str(row[idx["title_zh"]] or "").strip(),
                "content": str(row[idx[content_key]] or "").strip(),
                "cover": str(row[idx["cover_image"]] or "").strip(),
                "gallery": gallery,
            }
        )
    return rows


def detect_prefix_number(name: str, label: str) -> str | None:
    match = re.match(rf"^{label}(\d+)_", name)
    return match.group(1) if match else None


def find_matching_folder(base: Path, label: str, source_id: str, title: str) -> Path | None:
    suffix_num = source_id.split("-")[-1]
    direct = base / f"{label}{suffix_num}_{title}"
    if direct.exists():
        return direct
    for child in base.iterdir():
        if not child.is_dir():
            continue
        if child.name.startswith("99_") or child.name.startswith("补充_"):
            continue
        if title in child.name:
            return child
    return None


def rename_files_in_folder(folder: Path, old_label: str, new_label: str, year: str) -> None:
    for item in list(folder.iterdir()):
        if not item.is_file():
            continue
        old_name = item.name
        if old_name.startswith(old_label):
            tail = old_name[len(old_label):]
            new_name = f"{new_label}_{year}_{tail}"
            if new_name != old_name:
                item.rename(item.with_name(new_name))


def rename_category(base: Path, label: str, records: list[dict], sequential: bool) -> dict[str, Path]:
    mapping: dict[str, Path] = {}
    for index, record in enumerate(records, start=1):
        display_num = f"{index:03d}" if sequential else record["source_id"].split("-")[-1]
        old_folder = find_matching_folder(base, label, record["source_id"], record["title"])
        if old_folder is None:
            continue
        new_folder_name = f"{label}{display_num}_{record['year']}_{record['title']}"
        new_folder = old_folder.with_name(new_folder_name)
        if old_folder != new_folder and not new_folder.exists():
            old_folder.rename(new_folder)
            old_folder = new_folder
        old_label = f"{label}{detect_prefix_number(old_folder.name, label) or display_num}"
        new_label = f"{label}{display_num}"
        rename_files_in_folder(old_folder, old_label, new_label, record["year"])
        mapping[record["source_id"]] = old_folder
    return mapping


def image_paths(folder: Path) -> list[str]:
    allowed = {".webp", ".jpg", ".jpeg", ".png", ".mp4", ".mpg", ".wmv"}
    files = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in allowed]
    files.sort(key=lambda p: p.name)
    return [str(p.relative_to(PACK_ROOT)).replace("\\", "/") for p in files]


def rebuild_index(exhibitions: list[dict], news: list[dict], folder_maps: dict[str, dict[str, Path]]) -> None:
    workbook_path = PACK_ROOT / "张满堂资料查阅总表.xlsx"
    existing = load_workbook(workbook_path)
    for name in list(existing.sheetnames):
        if name in {"展览", "新闻"}:
            del existing[name]

    for sheet_name, records in [("展览", exhibitions), ("新闻", news)]:
        ws = existing.create_sheet(sheet_name)
        ws.append(["编号", "中文标题", "中文内容", "图片所在位置"])
        for index, record in enumerate(records, start=1):
            folder = folder_maps["展览" if sheet_name == "展览" else "新闻"].get(record["source_id"])
            display_id = f"{sheet_name}{index:03d}"
            title = f"{record['year']}_{record['title']}"
            ws.append([display_id, title, record["content"], "；".join(image_paths(folder)) if folder else ""])

    notes = existing["说明"] if "说明" in existing.sheetnames else existing.create_sheet("说明")
    notes["A2"] = "展览和新闻目录、资料卡、图片文件名已统一加上年份。"
    notes["A3"] = "新闻编号已按当前实际条目顺排，消除了跳号。"

    existing.save(workbook_path)


def main() -> None:
    exhibitions = load_records("exhibitions", "year", "description_zh")
    news = load_records("news", "date", "content_zh")
    exhibition_base = PACK_ROOT / "04_展览资料"
    news_base = PACK_ROOT / "05_新闻动态资料"

    exhibition_map = rename_category(exhibition_base, "展览", exhibitions, sequential=False)
    news_map = rename_category(news_base, "新闻", news, sequential=True)
    rebuild_index(exhibitions, news, {"展览": exhibition_map, "新闻": news_map})
    print(str(PACK_ROOT).encode("unicode_escape").decode("ascii"))


if __name__ == "__main__":
    main()
