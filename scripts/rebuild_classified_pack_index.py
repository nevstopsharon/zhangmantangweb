from __future__ import annotations

from pathlib import Path
import re

from openpyxl import Workbook, load_workbook


PACK_ROOT = Path(r"C:\Users\weixi\Desktop\张满堂资料总库_分类版")
SOURCE_WORKBOOK = Path(r"C:\Users\weixi\Desktop\zhangmantang-website\excel\content.xlsx")

BASE_MAP = {
    "works": PACK_ROOT / "03_艺术品资料",
    "exhibitions": PACK_ROOT / "04_展览资料",
    "news": PACK_ROOT / "05_新闻动态资料",
}
CONTENT_FIELD = {
    "works": "description_zh",
    "exhibitions": "description_zh",
    "news": "content_zh",
}
SHEET_NAME_MAP = {
    "works": "艺术品",
    "exhibitions": "展览",
    "news": "新闻",
}
LABEL_MAP = {
    "works": "作品",
    "exhibitions": "展览",
    "news": "新闻",
}
ALLOWED = {".webp", ".jpg", ".jpeg", ".png", ".mp4", ".mpg", ".wmv"}


def load_records(sheet: str) -> list[dict[str, str]]:
    wb = load_workbook(SOURCE_WORKBOOK, data_only=True)
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx["id"]]:
            continue
        rows.append(
            {
                "id": str(row[idx["id"]]),
                "title": str(row[idx["title_zh"]] or "").strip(),
                "year": str(row[idx["year" if sheet != "news" else "date"]] or "").strip(),
                "content": str(row[idx[CONTENT_FIELD[sheet]]] or "").strip(),
            }
        )
    return rows


def sorted_dirs(base: Path, label: str) -> list[Path]:
    dirs = [
        p for p in base.iterdir()
        if p.is_dir() and not p.name.startswith("99_") and not p.name.startswith("补充_")
    ]

    def key_fn(path: Path) -> int:
        match = re.match(rf"^{label}(\d+)", path.name)
        return int(match.group(1)) if match else 999999

    return sorted(dirs, key=key_fn)


def build_paths(folder: Path) -> str:
    files = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in ALLOWED]
    files.sort(key=lambda p: p.name)
    return "；".join(str(p.relative_to(PACK_ROOT)).replace("\\", "/") for p in files)


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet in ("works", "exhibitions", "news"):
        records = load_records(sheet)
        base = BASE_MAP[sheet]
        label = LABEL_MAP[sheet]
        dirs = sorted_dirs(base, label)
        ws = wb.create_sheet(SHEET_NAME_MAP[sheet])
        ws.append(["编号", "中文标题", "中文内容", "图片所在位置"])

        if sheet == "works":
            folder_by_title = {}
            for folder in dirs:
                parts = folder.name.split("_", 1)
                if len(parts) == 2:
                    folder_by_title[parts[1]] = folder
            for record in records:
                folder = folder_by_title.get(record["title"])
                ws.append([
                    record["id"],
                    record["title"],
                    record["content"],
                    build_paths(folder) if folder else "",
                ])
        elif sheet == "exhibitions":
            if len(dirs) != len(records):
                raise RuntimeError(f"展览目录数和记录数不一致: {len(dirs)} vs {len(records)}")
            for record, folder in zip(records, dirs):
                title = folder.name.split("_", 2)[2] if folder.name.count("_") >= 2 else record["title"]
                ws.append([
                    f"展览{record['id'].split('-')[-1]}",
                    f"{record['year']}_{title}",
                    record["content"],
                    build_paths(folder),
                ])
        else:
            if len(dirs) != len(records):
                raise RuntimeError(f"新闻目录数和记录数不一致: {len(dirs)} vs {len(records)}")
            for index, (record, folder) in enumerate(zip(records, dirs), start=1):
                title = folder.name.split("_", 2)[2] if folder.name.count("_") >= 2 else record["title"]
                ws.append([
                    f"新闻{index:03d}",
                    f"{record['year']}_{title}",
                    record["content"],
                    build_paths(folder),
                ])

    notes = wb.create_sheet("说明")
    notes.append(["说明"])
    notes.append(["展览和新闻目录、封面、图片、资料卡已统一加上年份。"])
    notes.append(["新闻编号已按当前条目顺排，不再缺号。"])
    notes.append(["图片所在位置已按当前分类版目录中的真实文件位置重建。"])

    wb.save(PACK_ROOT / "张满堂资料查阅总表.xlsx")
    print("done")


if __name__ == "__main__":
    main()
