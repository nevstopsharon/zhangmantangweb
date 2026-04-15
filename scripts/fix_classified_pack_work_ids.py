from __future__ import annotations

from pathlib import Path
import re

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\weixi\Desktop\张满堂资料总库_分类版\张满堂资料查阅总表.xlsx")


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["艺术品"]
    for row in ws.iter_rows(min_row=2):
        cell = row[0]
        value = str(cell.value or "").strip()
        match = re.fullmatch(r"work-(\d+)", value)
        if match:
            cell.value = f"作品{match.group(1)}"
    wb.save(WORKBOOK_PATH)


if __name__ == "__main__":
    main()
