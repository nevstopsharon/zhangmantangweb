from __future__ import annotations

import json
import re
import shutil
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(r"C:\Users\weixi\Desktop\zhangmantang-website")
SUPPLEMENT_ROOT = Path(r"C:\Users\weixi\Desktop\网站全部资料")
DESKTOP = Path(r"C:\Users\weixi\Desktop")

DESTINATION_NAME = f"\u5f20\u6ee1\u5802\u8d44\u6599\u603b\u5e93_\u7eaf\u8d44\u6599\u5ba2\u6237\u7248_{date.today().isoformat()}"
DESTINATION_ROOT = DESKTOP / DESTINATION_NAME

WORKBOOK_PATH = ROOT / "excel" / "content.xlsx"
PROFILE_PATH = ROOT / "data" / "profile.json"

INVALID_CHARS = re.compile(r'[<>:"/\\|?*]+')


SUPPLEMENT_MAP = {
    "news": {
        "news-002": [SUPPLEMENT_ROOT / "04_\u65b0\u95fb\u52a8\u6001_Credentials" / "\u4e2d\u5916\u753b\u520a\u4e13\u8bbf"],
        "news-003": [SUPPLEMENT_ROOT / "04_\u65b0\u95fb\u52a8\u6001_Credentials" / "2011 \u963f\u62c9\u4f2f\u4ea4\u6d41\u4f1a"],
        "news-004": [SUPPLEMENT_ROOT / "04_\u65b0\u95fb\u52a8\u6001_Credentials" / "06_2016\u5929\u5bab\u4e8c\u53f7\u592a\u7a7a\u827a\u672f"],
        "news-005": [SUPPLEMENT_ROOT / "04_\u65b0\u95fb\u52a8\u6001_Credentials" / "05_2023\u4e03\u56fd\u4e00\u5e26\u4e00\u8def\u90ae\u7968"],
        "news-006": [SUPPLEMENT_ROOT / "04_\u65b0\u95fb\u52a8\u6001_Credentials" / "04_2024\u4eba\u6c11\u5927\u4f1a\u5802\u6536\u85cf\u4f5c\u54c1"],
        "news-007": [SUPPLEMENT_ROOT / "04_\u65b0\u95fb\u52a8\u6001_Credentials" / "03_\u975e\u9057\u4f20\u627f\u4eba\u8ba4\u5b9a"],
        "news-008": [SUPPLEMENT_ROOT / "04_\u65b0\u95fb\u52a8\u6001_Credentials" / "2025\u8bfb\u8005\u6742\u5fd7"],
        "news-009": [SUPPLEMENT_ROOT / "04_\u65b0\u95fb\u52a8\u6001_Credentials" / "2026\u4e2d\u534e\u82f1\u624d"],
        "news-010": [SUPPLEMENT_ROOT / "04_\u65b0\u95fb\u52a8\u6001_Credentials" / "2024 \u5bf9\u8bdd\u6ee8\u5dde"],
        "news-011": [SUPPLEMENT_ROOT / "04_\u65b0\u95fb\u52a8\u6001_Credentials" / "2016 \u5927\u7231\u65e0\u8a00"],
    },
    "exhibitions": {
        "exhibition-001": [SUPPLEMENT_ROOT / "05_\u5c55\u89c8_Exhibitions" / "2007_\u519b\u535a114\u5468\u5e74\u5c55_114\u4f5c\u54c1"],
        "exhibition-002": [
            SUPPLEMENT_ROOT / "05_\u5c55\u89c8_Exhibitions" / "2007.12_\u656c\u4e34\u6bdb\u6cfd\u4e1c\u4e3b\u5e2d\u4e66\u6cd5\u4f5c\u54c1\u6c47\u62a5\u5c55",
            SUPPLEMENT_ROOT / "05_\u5c55\u89c8_Exhibitions" / "2008_\u656c\u4e34\u6bdb\u6cfd\u4e1c\u4e3b\u5e2d\u4e66\u6cd5\u4f5c\u54c1\u6c47\u62a5\u5c55",
        ],
        "exhibition-003": [SUPPLEMENT_ROOT / "05_\u5c55\u89c8_Exhibitions" / "2009_\u4e1c\u4eac\u4e2d\u56fd\u4e66\u6cd5\u7f8e\u672f\u4f5c\u54c1\u5c55"],
        "exhibition-004": [SUPPLEMENT_ROOT / "05_\u5c55\u89c8_Exhibitions" / "2010_\u4e0a\u6d77\u4e16\u535a\u671f\u95f4\u6c5f\u5c71\u5982\u6b64\u591a\u5a07\u7ea2\u8272\u4e66\u753b\u5c55_89\u4f5c\u54c1"],
        "exhibition-005": [SUPPLEMENT_ROOT / "05_\u5c55\u89c8_Exhibitions" / "2011_\u300a\u6bdb\u6cfd\u4e1c\u8bd7\u8bcd\u58a8\u5b9d\u5c55\u300b"],
        "exhibition-006": [SUPPLEMENT_ROOT / "05_\u5c55\u89c8_Exhibitions" / "2012_\u7eaa\u5ff5\u4e2d\u65b0\u5efa\u4ea440\u5468\u5e74\u4e66\u753b\u5c55"],
        "exhibition-007": [SUPPLEMENT_ROOT / "05_\u5c55\u89c8_Exhibitions" / "2013_\u4e2d\u56fd\u7b2c\u5341\u5c4a\u827a\u672f\u8282\u53e4\u4eca\u4e66\u6cd5\u827a\u672f\u5c55"],
        "exhibition-008": [SUPPLEMENT_ROOT / "05_\u5c55\u89c8_Exhibitions" / "2013_\u5317\u4eac\u519b\u535a\u767e\u540d\u5c06\u519b\u4e66\u753b\u5c55_120\u4f5c\u54c1"],
        "exhibition-009": [SUPPLEMENT_ROOT / "04_\u65b0\u95fb\u52a8\u6001_Credentials" / "01_2025\u5362\u6d6e\u5bab\u5c55\u89c8"],
    },
}


@dataclass
class ItemRecord:
    category: str
    item_id: str
    title_zh: str
    title_en: str
    fields: dict[str, str]
    cover: str
    gallery: list[str]


def safe_name(value: str) -> str:
    cleaned = INVALID_CHARS.sub(" ", str(value or "").strip())
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip().rstrip(".")


def ensure_unique_dir(path: Path) -> Path:
    if not path.exists():
        return path
    suffix = 2
    while True:
        candidate = path.with_name(f"{path.name}_{suffix}")
        if not candidate.exists():
            return candidate
        suffix += 1


def copy_file(src: Path, dest: Path) -> Path | None:
    if not src.exists() or not src.is_file():
        return None
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dest)
    return dest


def copy_tree(src: Path, dest: Path) -> bool:
    if not src.exists():
        return False
    shutil.copytree(src, dest, dirs_exist_ok=True)
    return True


def source_path_from_web(path_str: str) -> Path:
    normalized = path_str.lstrip("/").replace("/", "\\")
    return ROOT / normalized


def load_items() -> dict[str, list[ItemRecord]]:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    records: dict[str, list[ItemRecord]] = defaultdict(list)
    field_sets = {
        "works": [
            "year",
            "material_zh",
            "project_zh",
            "size_zh",
            "location_zh",
            "description_zh",
            "inspiration_zh",
        ],
        "exhibitions": ["year", "location_zh", "description_zh"],
        "news": [
            "date",
            "content_zh",
            "related_report_label_zh",
            "related_report_url",
            "full_video_label_zh",
            "full_video_url",
        ],
    }
    for sheet in ("works", "exhibitions", "news"):
        ws = wb[sheet]
        headers = [cell.value for cell in ws[1]]
        header_index = {header: idx for idx, header in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            item_id = row[header_index["id"]]
            if not item_id:
                continue
            gallery_raw = row[header_index["gallery_images"]] or ""
            gallery = [part.strip() for part in str(gallery_raw).split(";") if str(part).strip()]
            fields = {field: str(row[header_index[field]] or "").strip() for field in field_sets[sheet]}
            records[sheet].append(
                ItemRecord(
                    category=sheet,
                    item_id=str(item_id),
                    title_zh=str(row[header_index["title_zh"]] or "").strip(),
                    title_en=str(row[header_index["title_en"]] or "").strip(),
                    fields=fields,
                    cover=str(row[header_index["cover_image"]] or "").strip(),
                    gallery=gallery,
                )
            )
    return records


def build_item_folder_name(record: ItemRecord) -> str:
    return f"{record.item_id}_{safe_name(record.title_zh)}"


def write_metadata(record: ItemRecord, folder: Path, copied_files: list[str], supplements: list[str]) -> None:
    labels = {
        "works": {
            "year": "\u521b\u4f5c\u65f6\u95f4",
            "material_zh": "\u6750\u8d28",
            "project_zh": "\u9879\u76ee\u5f52\u5c5e",
            "size_zh": "\u5c3a\u5bf8",
            "location_zh": "\u5730\u70b9",
            "description_zh": "\u4f5c\u54c1\u8bf4\u660e",
            "inspiration_zh": "\u521b\u4f5c\u7075\u611f",
        },
        "exhibitions": {
            "year": "\u65f6\u95f4",
            "location_zh": "\u5730\u70b9",
            "description_zh": "\u5c55\u89c8\u8bf4\u660e",
        },
        "news": {
            "date": "\u65f6\u95f4",
            "content_zh": "\u65b0\u95fb\u5185\u5bb9",
            "related_report_label_zh": "\u76f8\u5173\u62a5\u9053\u6807\u7b7e",
            "related_report_url": "\u76f8\u5173\u62a5\u9053\u94fe\u63a5",
            "full_video_label_zh": "\u5b8c\u6574\u89c6\u9891\u6807\u7b7e",
            "full_video_url": "\u5b8c\u6574\u89c6\u9891\u94fe\u63a5",
        },
    }
    lines = [
        f"\u7f16\u53f7\uff1a{record.item_id}",
        f"\u4e2d\u6587\u6807\u9898\uff1a{record.title_zh}",
    ]
    for key, value in record.fields.items():
        if value:
            lines.append(f"{labels[record.category][key]}\uff1a{value}")
    lines.append("")
    lines.append("\u5f53\u524d\u6587\u4ef6\u5939\u5185\u590d\u5236\u6587\u4ef6\uff1a")
    lines.extend([f"- {name}" for name in copied_files] or ["- \u65e0"])
    lines.append("")
    lines.append("\u8865\u5145\u539f\u59cb\u8d44\u6599\u6765\u6e90\uff1a")
    lines.extend([f"- {entry}" for entry in supplements] or ["- \u65e0"])
    meta_name = f"{record.item_id}_{safe_name(record.title_zh)}_\u8d44\u6599\u5361.txt"
    (folder / meta_name).write_text("\n".join(lines), encoding="utf-8")


def copy_item_assets(record: ItemRecord, category_root: Path) -> tuple[Path, list[str], list[str]]:
    folder = category_root / build_item_folder_name(record)
    folder.mkdir(parents=True, exist_ok=True)
    copied_names: list[str] = []
    supplement_names: list[str] = []

    ordered_paths: list[tuple[str, str]] = []
    if record.cover:
        ordered_paths.append(("\u5c01\u9762", record.cover))
    ordered_paths.extend((f"\u56fe\u7247{index:02d}", path) for index, path in enumerate(record.gallery, start=1))

    for label, path_str in ordered_paths:
        src = source_path_from_web(path_str)
        if not src.exists():
            continue
        dest = folder / f"{record.item_id}_{safe_name(record.title_zh)}_{label}{src.suffix.lower()}"
        copied = copy_file(src, dest)
        if copied:
            copied_names.append(copied.name)

    for supplement_src in SUPPLEMENT_MAP.get(record.category, {}).get(record.item_id, []):
        if not supplement_src.exists():
            continue
        supplement_dest = folder / "\u8865\u5145\u539f\u59cb\u8d44\u6599" / safe_name(supplement_src.name)
        copy_tree(supplement_src, supplement_dest)
        supplement_names.append(str(supplement_src))

    write_metadata(record, folder, copied_names, supplement_names)
    return folder, copied_names, supplement_names


def build_profile_sections(dest_root: Path) -> None:
    profile = json.loads(PROFILE_PATH.read_text(encoding="utf-8"))
    home_dir = dest_root / "02_\u9996\u9875\u4e0e\u54c1\u724c\u8d44\u6599"
    about_dir = dest_root / "03_\u4f5c\u8005\u4ecb\u7ecd\u8d44\u6599"
    contact_dir = dest_root / "07_\u8054\u7cfb\u8d44\u6599"
    for directory in (home_dir, about_dir, contact_dir):
        directory.mkdir(parents=True, exist_ok=True)

    for path_str in [
        profile["brand"]["seal_image"],
        profile["brand"]["signature_image"],
        profile["hero"]["background_image"],
    ]:
        src = source_path_from_web(path_str)
        if src.exists():
            copy_file(src, home_dir / src.name)

    home_summary = [
        f"\u827a\u672f\u5bb6\u4e2d\u6587\u540d\uff1a{profile.get('artist_name_zh', '')}",
        f"\u827a\u672f\u5bb6\u82f1\u6587\u540d\uff1a{profile.get('artist_name_en', '')}",
        f"\u9996\u9875\u89c6\u89c9\u56fe\uff1a{profile.get('hero', {}).get('background_image', '')}",
    ]
    (home_dir / "\u9996\u9875\u4e0e\u54c1\u724c\u8d44\u6599\u8bf4\u660e.txt").write_text("\n".join(home_summary), encoding="utf-8")

    for index, path_str in enumerate(profile.get("about", {}).get("portrait_images", []), start=1):
        src = source_path_from_web(path_str)
        if src.exists():
            copy_file(src, about_dir / f"about_\u4eba\u7269\u7167\u7247_{index:02d}{src.suffix.lower()}")
    about_text = [
        f"\u5173\u4e8e\u9875\u6807\u9898\uff1a{profile.get('about', {}).get('headline_zh', '')}",
        "",
        "\u4eba\u7269\u4ecb\u7ecd\uff1a",
        profile.get("about", {}).get("bio_zh", ""),
    ]
    (about_dir / "\u4f5c\u8005\u4ecb\u7ecd\u8d44\u6599\u8bf4\u660e.txt").write_text("\n".join(about_text), encoding="utf-8")

    contact_text = [
        f"\u57ce\u5e02\uff1a{profile.get('contact', {}).get('city_zh', '')}",
        f"\u90ae\u7bb1\uff1a{profile.get('contact', {}).get('email', '')}",
        f"\u56de\u590d\u8bf4\u660e\uff1a{profile.get('contact', {}).get('reply_zh', '')}",
        f"Instagram\uff1a{profile.get('contact', {}).get('instagram', '')}",
    ]
    (contact_dir / "\u8054\u7cfb\u8d44\u6599\u8bf4\u660e.txt").write_text("\n".join(contact_text), encoding="utf-8")


def copy_full_supplement_backup(dest_root: Path) -> None:
    copy_tree(SUPPLEMENT_ROOT, dest_root / "90_\u8865\u5145\u539f\u59cb\u8d44\u6599_\u7f51\u7ad9\u5168\u90e8\u8d44\u6599\u539f\u6837\u5907\u4efd")


def create_index_workbook(dest_root: Path, records: dict[str, list[ItemRecord]], copy_log: dict[str, list[dict]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("\u603b\u89c8")
    summary.append(["\u5206\u7c7b", "\u6761\u76ee\u6570\u91cf"])
    summary.append(["\u827a\u672f\u54c1", len(records["works"])])
    summary.append(["\u5c55\u89c8", len(records["exhibitions"])])
    summary.append(["\u65b0\u95fb", len(records["news"])])

    headers = ["\u7f16\u53f7", "\u4e2d\u6587\u6807\u9898", "\u76ee\u6807\u6587\u4ef6\u5939", "\u590d\u5236\u6587\u4ef6\u6570\u91cf", "\u8865\u5145\u8d44\u6599\u6570\u91cf"]
    sheet_names = {
        "works": "\u827a\u672f\u54c1",
        "exhibitions": "\u5c55\u89c8",
        "news": "\u65b0\u95fb",
    }
    for key in ("works", "exhibitions", "news"):
        ws = wb.create_sheet(sheet_names[key])
        ws.append(headers)
        for item in copy_log[key]:
            ws.append([item["id"], item["title"], item["folder"], item["copied_count"], item["supplement_count"]])

    notes = wb.create_sheet("\u8bf4\u660e")
    notes.append(["\u8bf4\u660e"])
    notes.append(["\u672c\u7d22\u5f15\u7528\u4e8e\u5e2e\u52a9\u5ba2\u6237\u6309\u7f16\u53f7\u548c\u4e2d\u6587\u6807\u9898\u67e5\u9605\u8d44\u6599\u3002"])
    notes.append(["\u672c\u5305\u4ec5\u5305\u542b\u8d44\u6599\u526f\u672c\uff0c\u4e0d\u5305\u542b\u7f51\u7ad9\u4ee3\u7801\u3001\u811a\u672c\u3001\u6837\u5f0f\u6216\u9875\u9762\u8fd0\u884c\u6587\u4ef6\u3002"])
    notes.append(["\u5982\u9700\u67e5\u770b\u672a\u5f52\u6863\u7684\u8865\u5145\u539f\u59cb\u8d44\u6599\uff0c\u8bf7\u6253\u5f00 90_\u8865\u5145\u539f\u59cb\u8d44\u6599_\u7f51\u7ad9\u5168\u90e8\u8d44\u6599\u539f\u6837\u5907\u4efd\u3002"])

    index_dir = dest_root / "09_\u67e5\u9605\u7d22\u5f15"
    index_dir.mkdir(parents=True, exist_ok=True)
    wb.save(index_dir / "\u5f20\u6ee1\u5802\u8d44\u6599\u7d22\u5f15_\u7eaf\u8d44\u6599\u5ba2\u6237\u7248.xlsx")
    copy_file(WORKBOOK_PATH, index_dir / "\u7f51\u7ad9\u5185\u5bb9\u603b\u8868_\u53c2\u8003.xlsx")


def write_root_readme(dest_root: Path) -> None:
    text = """\u5f20\u6ee1\u5802\u8d44\u6599\u603b\u5e93\uff08\u7eaf\u8d44\u6599\u5ba2\u6237\u7248\uff09\u4f7f\u7528\u8bf4\u660e

1. \u672c\u76ee\u5f55\u4ec5\u5305\u542b\u4f9b\u67e5\u9605\u7684\u8d44\u6599\u526f\u672c\uff0c\u4e0d\u5305\u542b\u4efb\u4f55\u7f51\u7ad9\u4ee3\u7801\u3001\u524d\u7aef\u811a\u672c\u3001\u6837\u5f0f\u6216\u53ef\u8fd0\u884c\u9875\u9762\u3002

2. 02_\u9996\u9875\u4e0e\u54c1\u724c\u8d44\u6599\u300103_\u4f5c\u8005\u4ecb\u7ecd\u8d44\u6599\u300107_\u8054\u7cfb\u8d44\u6599
   \u5b58\u653e\u9996\u9875\u3001\u4f5c\u8005\u4ecb\u7ecd\u3001\u8054\u7cfb\u4fe1\u606f\u7b49\u57fa\u7840\u8d44\u6599\u3002

3. 04_\u827a\u672f\u54c1\u8d44\u6599\u300105_\u5c55\u89c8\u8d44\u6599\u300106_\u65b0\u95fb\u52a8\u6001\u8d44\u6599
   \u6bcf\u4e2a\u6761\u76ee\u5747\u6309\u201c\u7f16\u53f7 + \u4e2d\u6587\u6807\u9898\u201d\u547d\u540d\u3002
   \u6587\u4ef6\u5939\u5185\u5305\u542b\uff1a
   - \u8d44\u6599\u5361
   - \u7f51\u7ad9\u4e2d\u5df2\u6574\u7406\u597d\u7684\u56fe\u7247\u6216\u89c6\u9891\u526f\u672c
   - \u80fd\u5bf9\u5e94\u4e0a\u7684\u8865\u5145\u539f\u59cb\u8d44\u6599

4. 09_\u67e5\u9605\u7d22\u5f15
   \u63d0\u4f9b Excel \u7d22\u5f15\u548c\u7f51\u7ad9\u5185\u5bb9\u603b\u8868\u53c2\u8003\uff0c\u65b9\u4fbf\u6309\u7f16\u53f7\u548c\u6807\u9898\u5feb\u901f\u67e5\u627e\u3002

5. 90_\u8865\u5145\u539f\u59cb\u8d44\u6599_\u7f51\u7ad9\u5168\u90e8\u8d44\u6599\u539f\u6837\u5907\u4efd
   \u5b58\u653e\u672a\u7ed3\u6784\u5316\u7684\u5b8c\u6574\u8865\u5145\u539f\u59cb\u8d44\u6599\u5907\u4efd\uff0c\u4fbf\u4e8e\u540e\u7eed\u7ee7\u7eed\u6574\u7406\u3002
"""
    readme_dir = dest_root / "00_\u8bf7\u5148\u9605\u8bfb"
    readme_dir.mkdir(parents=True, exist_ok=True)
    (readme_dir / "\u5ba2\u6237\u7248\u8d44\u6599\u5e93\u4f7f\u7528\u8bf4\u660e.txt").write_text(text, encoding="utf-8")


def main() -> None:
    dest_root = ensure_unique_dir(DESTINATION_ROOT)
    dest_root.mkdir(parents=True, exist_ok=True)

    build_profile_sections(dest_root)
    records = load_items()
    category_dirs = {
        "works": dest_root / "04_\u827a\u672f\u54c1\u8d44\u6599",
        "exhibitions": dest_root / "05_\u5c55\u89c8\u8d44\u6599",
        "news": dest_root / "06_\u65b0\u95fb\u52a8\u6001\u8d44\u6599",
    }
    for path in category_dirs.values():
        path.mkdir(parents=True, exist_ok=True)

    copy_log: dict[str, list[dict]] = defaultdict(list)
    for category, items in records.items():
        for record in items:
            folder, copied_names, supplement_names = copy_item_assets(record, category_dirs[category])
            copy_log[category].append(
                {
                    "id": record.item_id,
                    "title": record.title_zh,
                    "folder": str(folder),
                    "copied_count": len(copied_names),
                    "supplement_count": len(supplement_names),
                }
            )

    qianlong_src = SUPPLEMENT_ROOT / "04_\u65b0\u95fb\u52a8\u6001_Credentials" / "\u5317\u4eac\u5343\u9f99\u7f51"
    if qianlong_src.exists():
        copy_tree(qianlong_src, dest_root / "06_\u65b0\u95fb\u52a8\u6001\u8d44\u6599" / "\u8865\u5145_\u5317\u4eac\u5343\u9f99\u7f51\u89c6\u9891")

    artwork_raw = SUPPLEMENT_ROOT / "03_\u827a\u672f\u54c1_Artwork"
    if artwork_raw.exists():
        copy_tree(artwork_raw, dest_root / "04_\u827a\u672f\u54c1\u8d44\u6599" / "99_\u8865\u5145\u539f\u59cb\u7d20\u6750_\u6765\u81ea\u7f51\u7ad9\u5168\u90e8\u8d44\u6599")

    copy_full_supplement_backup(dest_root)
    write_root_readme(dest_root)
    create_index_workbook(dest_root, records, copy_log)
    print(str(dest_root).encode("unicode_escape").decode("ascii"))


if __name__ == "__main__":
    main()
