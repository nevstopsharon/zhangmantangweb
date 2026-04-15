from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook


PACK_ROOT = Path(r"C:\Users\weixi\Desktop\张满堂资料总库_分类版")
SOURCE_WORKBOOK = Path(r"C:\Users\weixi\Desktop\zhangmantang-website\excel\content.xlsx")


PREFIX_MAP = {
    "works": ("作品", PACK_ROOT / "03_艺术品资料"),
    "exhibitions": ("展览", PACK_ROOT / "04_展览资料"),
    "news": ("新闻", PACK_ROOT / "05_新闻动态资料"),
}


def load_source_records():
    wb = load_workbook(SOURCE_WORKBOOK, data_only=True)
    result = {}
    content_field = {
        "works": "description_zh",
        "exhibitions": "description_zh",
        "news": "content_zh",
    }
    for sheet in ("works", "exhibitions", "news"):
        ws = wb[sheet]
        headers = [cell.value for cell in ws[1]]
        idx = {header: i for i, header in enumerate(headers)}
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item_id = row[idx["id"]]
            if not item_id:
                continue
            gallery_raw = str(row[idx["gallery_images"]] or "").strip()
            gallery = [part.strip() for part in gallery_raw.split(";") if part.strip()]
            rows.append(
                {
                    "id": str(item_id),
                    "title": str(row[idx["title_zh"]] or "").strip(),
                    "content": str(row[idx[content_field[sheet]]] or "").strip(),
                    "cover": str(row[idx["cover_image"]] or "").strip(),
                    "gallery": gallery,
                }
            )
        result[sheet] = rows
    return result


def rename_category_folders(category: str, label: str, root: Path) -> dict[str, Path]:
    mapping: dict[str, Path] = {}
    if not root.exists():
        return mapping
    pattern = re.compile(rf"^{category[:-1] if category.endswith('s') else category}-(\d+?)_(.+)$")
    for child in root.iterdir():
        if not child.is_dir():
            continue
        if child.name.startswith("99_") or child.name.startswith("补充_"):
            continue
        if category == "works":
            pattern = re.compile(r"^work-(\d+)_(.+)$")
        elif category == "exhibitions":
            pattern = re.compile(r"^exhibition-(\d+)_(.+)$")
        else:
            pattern = re.compile(r"^news-(\d+)_(.+)$")
        match = pattern.match(child.name)
        if not match:
            parts = child.name.split("_", 1)
            if len(parts) == 2:
                mapping[parts[0]] = child
            continue
        number, title = match.groups()
        new_name = f"{label}{number}_{title}"
        target = child.with_name(new_name)
        if target != child and not target.exists():
            child.rename(target)
            child = target
        mapping[number] = child
    return mapping


def build_image_paths(folder: Path) -> list[str]:
    files = [p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in {".webp", ".jpg", ".jpeg", ".png", ".mp4", ".mpg", ".wmv"}]
    files.sort(key=lambda p: p.name)
    return [str(p.relative_to(PACK_ROOT)).replace("\\", "/") for p in files]


def rebuild_index(records: dict, folder_maps: dict[str, dict[str, Path]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    sheet_names = {
        "works": "艺术品",
        "exhibitions": "展览",
        "news": "新闻",
    }
    for key in ("works", "exhibitions", "news"):
        ws = wb.create_sheet(sheet_names[key])
        ws.append(["编号", "中文标题", "中文内容", "图片所在位置"])
        for record in records[key]:
            number = record["id"].split("-")[-1]
            folder = folder_maps[key].get(number)
            image_paths = build_image_paths(folder) if folder and folder.exists() else []
            ws.append([record["id"], record["title"], record["content"], "；".join(image_paths)])

    notes = wb.create_sheet("说明")
    notes.append(["说明"])
    notes.append(["本表图片所在位置已按当前分类版资料夹中的实际文件位置更新。"])
    notes.append(["子文件夹已统一改为中文前缀：作品、展览、新闻。"])

    out_path = PACK_ROOT / "张满堂资料查阅总表.xlsx"
    wb.save(out_path)


def main() -> None:
    records = load_source_records()
    folder_maps = {}
    folder_maps["works"] = rename_category_folders("works", "作品", PREFIX_MAP["works"][1])
    folder_maps["exhibitions"] = rename_category_folders("exhibitions", "展览", PREFIX_MAP["exhibitions"][1])
    folder_maps["news"] = rename_category_folders("news", "新闻", PREFIX_MAP["news"][1])
    rebuild_index(records, folder_maps)
    print(str(PACK_ROOT).encode("unicode_escape").decode("ascii"))


if __name__ == "__main__":
    main()
