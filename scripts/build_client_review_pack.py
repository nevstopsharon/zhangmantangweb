from __future__ import annotations

import json
import re
import shutil
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(r"C:\Users\weixi\Desktop\zhangmantang-website")
SUPPLEMENT_ROOT = Path(r"C:\Users\weixi\Desktop\网站全部资料")
DESKTOP = Path(r"C:\Users\weixi\Desktop")

DESTINATION_NAME = f"张满堂资料总库_客户查阅版_{date.today().isoformat()}"
DESTINATION_ROOT = DESKTOP / DESTINATION_NAME

WORKBOOK_PATH = ROOT / "excel" / "content.xlsx"
PROFILE_PATH = ROOT / "data" / "profile.json"

INVALID_CHARS = re.compile(r'[<>:"/\\|?*]+')


SUPPLEMENT_MAP = {
    "news": {
        "news-002": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "中外画刊专访"],
        "news-003": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "2011 阿拉伯交流会"],
        "news-004": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "06_2016天宫二号太空艺术"],
        "news-005": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "05_2023七国一带一路邮票"],
        "news-006": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "04_2024人民大会堂收藏作品"],
        "news-007": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "03_非遗传承人认定"],
        "news-008": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "2025读者杂志"],
        "news-009": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "2026中华英才"],
        "news-010": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "2024 对话滨州"],
        "news-011": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "2016 大爱无言"],
    },
    "exhibitions": {
        "exhibition-001": [SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2007_军博114周年展_114作品"],
        "exhibition-002": [
            SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2007.12_敬临毛泽东主席书法作品汇报展",
            SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2008_敬临毛泽东主席书法作品汇报展",
        ],
        "exhibition-003": [SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2009_东京中国书法美术作品展"],
        "exhibition-004": [SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2010_上海世博期间江山如此多娇红色书画展_89作品"],
        "exhibition-005": [SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2011_《毛泽东诗词墨宝展》"],
        "exhibition-006": [SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2012_纪念中新建交40周年书画展"],
        "exhibition-007": [SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2013_中国第十届艺术节古今书法艺术展"],
        "exhibition-008": [SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2013_北京军博百名将军书画展_120作品"],
        "exhibition-009": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "01_2025卢浮宫展览"],
    },
}


@dataclass
class ItemRecord:
    category: str
    item_id: str
    title_zh: str
    content_zh: str
    cover: str
    gallery: list[str]


def safe_name(value: str) -> str:
    cleaned = INVALID_CHARS.sub(" ", str(value or "").strip())
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip().rstrip(".")


def ensure_unique_dir(path: Path) -> Path:
    if not path.exists():
        return path
    suffix = 2
    while True:
        candidate = path.with_name(f"{path.name}_{suffix}")
        if not candidate.exists():
            return candidate
        suffix += 1


def copy_file(src: Path, dest: Path) -> Path | None:
    if not src.exists() or not src.is_file():
        return None
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dest)
    return dest


def copy_tree(src: Path, dest: Path) -> bool:
    if not src.exists():
        return False
    shutil.copytree(src, dest, dirs_exist_ok=True)
    return True


def source_path_from_web(path_str: str) -> Path:
    normalized = path_str.lstrip("/").replace("/", "\\")
    return ROOT / normalized


def build_item_folder_name(record: ItemRecord) -> str:
    return f"{record.item_id}_{safe_name(record.title_zh)}"


def content_field_name(category: str) -> str:
    return {
        "works": "description_zh",
        "exhibitions": "description_zh",
        "news": "content_zh",
    }[category]


def load_items() -> dict[str, list[ItemRecord]]:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    records: dict[str, list[ItemRecord]] = defaultdict(list)
    for sheet in ("works", "exhibitions", "news"):
        ws = wb[sheet]
        headers = [cell.value for cell in ws[1]]
        header_index = {header: idx for idx, header in enumerate(headers)}
        content_key = content_field_name(sheet)
        for row in ws.iter_rows(min_row=2, values_only=True):
            item_id = row[header_index["id"]]
            if not item_id:
                continue
            gallery_raw = row[header_index["gallery_images"]] or ""
            gallery = [part.strip() for part in str(gallery_raw).split(";") if str(part).strip()]
            records[sheet].append(
                ItemRecord(
                    category=sheet,
                    item_id=str(item_id),
                    title_zh=str(row[header_index["title_zh"]] or "").strip(),
                    content_zh=str(row[header_index[content_key]] or "").strip(),
                    cover=str(row[header_index["cover_image"]] or "").strip(),
                    gallery=gallery,
                )
            )
    return records


def write_metadata(record: ItemRecord, folder: Path, copied_files: list[str], supplements: list[str]) -> None:
    lines = [
        f"编号：{record.item_id}",
        f"中文标题：{record.title_zh}",
        "",
        "中文内容：",
        record.content_zh or "无",
        "",
        "当前文件夹内资料：",
    ]
    lines.extend([f"- {name}" for name in copied_files] or ["- 无"])
    lines.append("")
    lines.append("补充原始资料来源：")
    lines.extend([f"- {entry}" for entry in supplements] or ["- 无"])
    meta_name = f"{record.item_id}_{safe_name(record.title_zh)}_资料卡.txt"
    (folder / meta_name).write_text("\n".join(lines), encoding="utf-8")


def copy_item_assets(record: ItemRecord, category_root: Path) -> tuple[Path, list[str], list[str], list[str]]:
    folder = category_root / build_item_folder_name(record)
    folder.mkdir(parents=True, exist_ok=True)
    copied_names: list[str] = []
    copied_relpaths: list[str] = []
    supplement_names: list[str] = []

    ordered_paths: list[tuple[str, str]] = []
    if record.cover:
        ordered_paths.append(("封面", record.cover))
    ordered_paths.extend((f"图片{index:02d}", path) for index, path in enumerate(record.gallery, start=1))

    for label, path_str in ordered_paths:
        src = source_path_from_web(path_str)
        if not src.exists():
            continue
        dest = folder / f"{record.item_id}_{safe_name(record.title_zh)}_{label}{src.suffix.lower()}"
        copied = copy_file(src, dest)
        if copied:
            copied_names.append(copied.name)
            copied_relpaths.append(str(copied.relative_to(folder.parent.parent.parent)).replace("\\", "/"))

    for supplement_src in SUPPLEMENT_MAP.get(record.category, {}).get(record.item_id, []):
        if not supplement_src.exists():
            continue
        supplement_dest = folder / "补充原始资料" / safe_name(supplement_src.name)
        copy_tree(supplement_src, supplement_dest)
        supplement_names.append(str(supplement_src))

    write_metadata(record, folder, copied_names, supplement_names)
    return folder, copied_names, supplement_names, copied_relpaths


def build_profile_sections(dest_root: Path) -> None:
    profile = json.loads(PROFILE_PATH.read_text(encoding="utf-8"))
    home_dir = dest_root / "01_首页与品牌资料"
    about_dir = dest_root / "02_作者介绍资料"
    contact_dir = dest_root / "06_联系资料"
    for directory in (home_dir, about_dir, contact_dir):
        directory.mkdir(parents=True, exist_ok=True)

    for path_str in [
        profile["brand"]["seal_image"],
        profile["brand"]["signature_image"],
        profile["hero"]["background_image"],
    ]:
        src = source_path_from_web(path_str)
        if src.exists():
            copy_file(src, home_dir / src.name)

    (home_dir / "首页与品牌资料说明.txt").write_text(
        "\n".join(
            [
                f"艺术家中文名：{profile.get('artist_name_zh', '')}",
                f"首页主视觉图：{profile.get('hero', {}).get('background_image', '')}",
            ]
        ),
        encoding="utf-8",
    )

    for index, path_str in enumerate(profile.get("about", {}).get("portrait_images", []), start=1):
        src = source_path_from_web(path_str)
        if src.exists():
            copy_file(src, about_dir / f"作者照片_{index:02d}{src.suffix.lower()}")
    (about_dir / "作者介绍资料说明.txt").write_text(
        "\n".join(
            [
                f"关于页标题：{profile.get('about', {}).get('headline_zh', '')}",
                "",
                "人物介绍：",
                profile.get("about", {}).get("bio_zh", ""),
            ]
        ),
        encoding="utf-8",
    )

    (contact_dir / "联系资料说明.txt").write_text(
        "\n".join(
            [
                f"城市：{profile.get('contact', {}).get('city_zh', '')}",
                f"邮箱：{profile.get('contact', {}).get('email', '')}",
                f"回复说明：{profile.get('contact', {}).get('reply_zh', '')}",
                f"Instagram：{profile.get('contact', {}).get('instagram', '')}",
            ]
        ),
        encoding="utf-8",
    )


def copy_full_supplement_backup(dest_root: Path) -> None:
    copy_tree(SUPPLEMENT_ROOT, dest_root / "90_补充原始资料备份")


def create_review_workbook(dest_root: Path, logs: dict[str, list[dict]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for key, title in (("works", "艺术品"), ("exhibitions", "展览"), ("news", "新闻")):
        ws = wb.create_sheet(title)
        ws.append(["编号", "中文标题", "中文内容", "图片所在位置"])
        for item in logs[key]:
            ws.append([item["id"], item["title"], item["content"], "；".join(item["image_paths"])])

    notes = wb.create_sheet("说明")
    notes.append(["说明"])
    notes.append(["本表仅保留客户查阅需要的四列：编号、中文标题、中文内容、图片所在位置。"])
    notes.append(["本资料包不包含任何网站代码、脚本、样式、页面文件或 JSON 文件。"])

    out_dir = dest_root / "07_查阅索引"
    out_dir.mkdir(parents=True, exist_ok=True)
    wb.save(out_dir / "张满堂资料查阅总表.xlsx")


def write_root_readme(dest_root: Path) -> None:
    text = """张满堂资料总库（客户查阅版）使用说明

1. 本目录只保留客户查阅资料所需要的内容。
2. 不包含任何网站代码、脚本、样式、页面文件或 JSON 文件。
3. 艺术品、展览、新闻都按“编号 + 中文标题”命名。
4. 如需总览，请打开 07_查阅索引\张满堂资料查阅总表.xlsx。
5. 如需查看未进一步整理的完整补充资料，请打开 90_补充原始资料备份。
"""
    readme_dir = dest_root / "00_请先阅读"
    readme_dir.mkdir(parents=True, exist_ok=True)
    (readme_dir / "客户资料库使用说明.txt").write_text(text, encoding="utf-8")


def main() -> None:
    dest_root = ensure_unique_dir(DESTINATION_ROOT)
    dest_root.mkdir(parents=True, exist_ok=True)

    build_profile_sections(dest_root)
    records = load_items()
    category_dirs = {
        "works": dest_root / "03_艺术品资料",
        "exhibitions": dest_root / "04_展览资料",
        "news": dest_root / "05_新闻动态资料",
    }
    for path in category_dirs.values():
        path.mkdir(parents=True, exist_ok=True)

    logs: dict[str, list[dict]] = defaultdict(list)
    for category, items in records.items():
        for record in items:
            folder, _copied_names, supplement_names, copied_relpaths = copy_item_assets(record, category_dirs[category])
            logs[category].append(
                {
                    "id": record.item_id,
                    "title": record.title_zh,
                    "content": record.content_zh,
                    "folder": str(folder),
                    "image_paths": copied_relpaths,
                    "supplement_count": len(supplement_names),
                }
            )

    qianlong_src = SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "北京千龙网"
    if qianlong_src.exists():
        copy_tree(qianlong_src, dest_root / "05_新闻动态资料" / "补充_北京千龙网视频")

    artwork_raw = SUPPLEMENT_ROOT / "03_艺术品_Artwork"
    if artwork_raw.exists():
        copy_tree(artwork_raw, dest_root / "03_艺术品资料" / "99_补充原始素材")

    copy_full_supplement_backup(dest_root)
    write_root_readme(dest_root)
    create_review_workbook(dest_root, logs)

    print(str(dest_root).encode("unicode_escape").decode("ascii"))


if __name__ == "__main__":
    main()
