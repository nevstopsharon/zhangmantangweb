from __future__ import annotations

import json
import re
import shutil
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(r"C:\Users\weixi\Desktop\zhangmantang-website")
SUPPLEMENT_ROOT = Path(r"C:\Users\weixi\Desktop\网站全部资料")
DESKTOP = Path(r"C:\Users\weixi\Desktop")


DESTINATION_NAME = f"张满堂资料总库_客户版_{date.today().isoformat()}"
DESTINATION_ROOT = DESKTOP / DESTINATION_NAME

WORKBOOK_PATH = ROOT / "excel" / "content.xlsx"
PROFILE_PATH = ROOT / "data" / "profile.json"


INVALID_CHARS = re.compile(r'[<>:"/\\\\|?*]+')


SUPPLEMENT_MAP = {
    "news": {
        "news-002": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "中外画刊专访"],
        "news-003": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "2011 阿拉伯交流会"],
        "news-004": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "06_2016天宫二号太空艺术"],
        "news-005": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "05_2023七国一带一路邮票"],
        "news-006": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "04_2024人民大会堂收藏作品"],
        "news-007": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "03_非遗传承人认定"],
        "news-008": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "2025读者杂志"],
        "news-009": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "2026中华英才"],
        "news-010": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "2024 对话滨州"],
        "news-011": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "2016 大爱无言"],
    },
    "exhibitions": {
        "exhibition-001": [SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2007_军博114周年展_114作品"],
        "exhibition-002": [
            SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2007.12_敬临毛泽东主席书法作品汇报展",
            SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2008_敬临毛泽东主席书法作品汇报展",
        ],
        "exhibition-003": [SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2009_东京中国书法美术作品展"],
        "exhibition-004": [SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2010_上海世博期间江山如此多娇红色书画展_89作品"],
        "exhibition-005": [SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2011_《毛泽东诗词墨宝展》"],
        "exhibition-006": [SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2012_纪念中新建交40周年书画展"],
        "exhibition-007": [SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2013_中国第十届艺术节古今书法艺术展"],
        "exhibition-008": [SUPPLEMENT_ROOT / "05_展览_Exhibitions" / "2013_北京军博百名将军书画展_120作品"],
        "exhibition-009": [SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "01_2025卢浮宫展览"],
    },
}


@dataclass
class ItemRecord:
    category: str
    item_id: str
    title_zh: str
    title_en: str
    fields: dict[str, str]
    cover: str
    gallery: list[str]


def safe_name(value: str) -> str:
    cleaned = INVALID_CHARS.sub(" ", str(value or "").strip())
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip().rstrip(".")


def ensure_unique_dir(path: Path) -> Path:
    if not path.exists():
        return path
    suffix = 2
    while True:
        candidate = path.with_name(f"{path.name}_{suffix}")
        if not candidate.exists():
            return candidate
        suffix += 1


def load_items() -> dict[str, list[ItemRecord]]:
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    records: dict[str, list[ItemRecord]] = defaultdict(list)
    field_sets = {
        "works": [
            "year",
            "material_zh",
            "project_zh",
            "size_zh",
            "location_zh",
            "description_zh",
            "inspiration_zh",
        ],
        "exhibitions": [
            "year",
            "location_zh",
            "description_zh",
        ],
        "news": [
            "date",
            "content_zh",
            "related_report_label_zh",
            "related_report_url",
            "full_video_label_zh",
            "full_video_url",
        ],
    }
    for sheet in ("works", "exhibitions", "news"):
        ws = wb[sheet]
        headers = [cell.value for cell in ws[1]]
        header_index = {header: idx for idx, header in enumerate(headers)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            item_id = row[header_index["id"]]
            if not item_id:
                continue
            gallery_raw = row[header_index["gallery_images"]] or ""
            gallery = [part.strip() for part in str(gallery_raw).split(";") if part and str(part).strip()]
            fields = {field: str(row[header_index[field]] or "").strip() for field in field_sets[sheet]}
            records[sheet].append(
                ItemRecord(
                    category=sheet,
                    item_id=str(item_id),
                    title_zh=str(row[header_index["title_zh"]] or "").strip(),
                    title_en=str(row[header_index["title_en"]] or "").strip(),
                    fields=fields,
                    cover=str(row[header_index["cover_image"]] or "").strip(),
                    gallery=gallery,
                )
            )
    return records


def copy_file(src: Path, dest: Path) -> Path | None:
    if not src.exists() or not src.is_file():
        return None
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dest)
    return dest


def copy_tree(src: Path, dest: Path) -> bool:
    if not src.exists():
        return False
    shutil.copytree(src, dest, dirs_exist_ok=True)
    return True


def source_path_from_web(path_str: str) -> Path:
    normalized = path_str.lstrip("/").replace("/", "\\")
    return ROOT / normalized


def build_item_folder_name(record: ItemRecord) -> str:
    return f"{record.item_id}_{safe_name(record.title_zh)}"


def write_metadata(record: ItemRecord, folder: Path, copied_files: list[str], supplements: list[str]) -> None:
    labels = {
        "works": {
            "year": "创作时间",
            "material_zh": "介质",
            "project_zh": "项目归属",
            "size_zh": "尺寸",
            "location_zh": "地点",
            "description_zh": "作品说明",
            "inspiration_zh": "创作灵感",
        },
        "exhibitions": {
            "year": "时间",
            "location_zh": "地点",
            "description_zh": "展览说明",
        },
        "news": {
            "date": "时间",
            "content_zh": "新闻内容",
            "related_report_label_zh": "相关报道标签",
            "related_report_url": "相关报道链接",
            "full_video_label_zh": "完整视频标签",
            "full_video_url": "完整视频链接",
        },
    }
    lines = [
        f"编号：{record.item_id}",
        f"中文标题：{record.title_zh}",
    ]
    if record.title_en:
        lines.append(f"英文标题：{record.title_en}")
    lines.append("")
    for key, value in record.fields.items():
        if value:
            lines.append(f"{labels[record.category][key]}：{value}")
    lines.append("")
    lines.append("当前文件夹内复制文件：")
    if copied_files:
        lines.extend([f"- {name}" for name in copied_files])
    else:
        lines.append("- 无")
    lines.append("")
    lines.append("补充原始资料来源：")
    if supplements:
        lines.extend([f"- {entry}" for entry in supplements])
    else:
        lines.append("- 无")
    meta_name = f"{record.item_id}_{safe_name(record.title_zh)}_资料卡.txt"
    (folder / meta_name).write_text("\n".join(lines), encoding="utf-8")


def copy_item_assets(record: ItemRecord, category_root: Path) -> tuple[Path, list[str], list[str]]:
    folder = category_root / build_item_folder_name(record)
    folder.mkdir(parents=True, exist_ok=True)
    copied_names: list[str] = []
    supplement_names: list[str] = []

    ordered_paths = []
    if record.cover:
        ordered_paths.append(("封面", record.cover))
    ordered_paths.extend((f"图片{index:02d}", path) for index, path in enumerate(record.gallery, start=1))

    for label, path_str in ordered_paths:
        src = source_path_from_web(path_str)
        if not src.exists():
            continue
        suffix = src.suffix.lower()
        base = f"{record.item_id}_{safe_name(record.title_zh)}_{label}{suffix}"
        dest = folder / base
        copied = copy_file(src, dest)
        if copied:
            copied_names.append(copied.name)

    for supplement_src in SUPPLEMENT_MAP.get(record.category, {}).get(record.item_id, []):
        if not supplement_src.exists():
            continue
        supplement_dest = folder / "补充原始资料" / safe_name(supplement_src.name)
        copy_tree(supplement_src, supplement_dest)
        supplement_names.append(str(supplement_src))

    write_metadata(record, folder, copied_names, supplement_names)
    return folder, copied_names, supplement_names


def build_profile_sections(dest_root: Path) -> None:
    profile = json.loads(PROFILE_PATH.read_text(encoding="utf-8"))

    home_dir = dest_root / "02_首页与品牌资料"
    about_dir = dest_root / "03_作者介绍资料"
    contact_dir = dest_root / "07_联系资料"
    docs_dir = dest_root / "08_项目文档与设计系统"
    for directory in (home_dir, about_dir, contact_dir, docs_dir):
        directory.mkdir(parents=True, exist_ok=True)

    for path_str in [profile["brand"]["seal_image"], profile["brand"]["signature_image"], profile["hero"]["background_image"]]:
        src = source_path_from_web(path_str)
        if src.exists():
            copy_file(src, home_dir / src.name)

    home_summary = [
        f"艺术家中文名：{profile.get('artist_name_zh', '')}",
        f"艺术家英文名：{profile.get('artist_name_en', '')}",
        f"首页视觉图：{profile.get('hero', {}).get('background_image', '')}",
    ]
    (home_dir / "首页与品牌资料说明.txt").write_text("\n".join(home_summary), encoding="utf-8")

    portrait_paths = profile.get("about", {}).get("portrait_images", [])
    for index, path_str in enumerate(portrait_paths, start=1):
        src = source_path_from_web(path_str)
        if src.exists():
            copy_file(src, about_dir / f"about_人物照片_{index:02d}{src.suffix.lower()}")
    about_text = [
        f"关于页标题：{profile.get('about', {}).get('headline_zh', '')}",
        "",
        "人物介绍：",
        profile.get("about", {}).get("bio_zh", ""),
    ]
    (about_dir / "作者介绍资料说明.txt").write_text("\n".join(about_text), encoding="utf-8")

    contact_text = [
        f"城市：{profile.get('contact', {}).get('city_zh', '')}",
        f"邮箱：{profile.get('contact', {}).get('email', '')}",
        f"回复说明：{profile.get('contact', {}).get('reply_zh', '')}",
        f"Instagram：{profile.get('contact', {}).get('instagram', '')}",
    ]
    (contact_dir / "联系资料说明.txt").write_text("\n".join(contact_text), encoding="utf-8")

    # Project documentation
    for file_name in [
        "README.md",
        "4.6 网站产品文档.docx",
        "requirements.txt",
    ]:
        src = ROOT / file_name
        if src.exists():
            copy_file(src, docs_dir / src.name)
    copy_tree(ROOT / "templates" / "design-system", docs_dir / "design-system")
    copy_file(WORKBOOK_PATH, docs_dir / "content.xlsx")


def build_site_runtime(dest_root: Path) -> None:
    runtime_dir = dest_root / "01_网站展示版"
    for name in ("site", "data", "images"):
        copy_tree(ROOT / name, runtime_dir / name)
    copy_file(WORKBOOK_PATH, runtime_dir / "excel" / "content.xlsx")
    for file_name in ("README.md", "4.6 网站产品文档.docx"):
        src = ROOT / file_name
        if src.exists():
            copy_file(src, runtime_dir / src.name)
    readme = (
        "这是当前网站展示版复制件。\n"
        "如需本地查看网站，请在该文件夹上级目录用 Python 本地服务打开。\n"
        "例如：py -m http.server 8000\n"
        "然后访问：http://127.0.0.1:8000/01_网站展示版/site/\n"
    )
    (runtime_dir / "使用说明.txt").write_text(readme, encoding="utf-8")


def copy_full_supplement_backup(dest_root: Path) -> None:
    copy_tree(SUPPLEMENT_ROOT, dest_root / "90_补充原始资料_网站全部资料原样备份")


def create_index_workbook(dest_root: Path, records: dict[str, list[ItemRecord]], copy_log: dict[str, list[dict]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("总览")
    summary.append(["分类", "条目数量"])
    summary.append(["艺术品", len(records["works"])])
    summary.append(["展览", len(records["exhibitions"])])
    summary.append(["新闻", len(records["news"])])

    headers = ["编号", "中文标题", "目标文件夹", "复制文件数量", "补充资料数量"]
    sheet_names = {
        "works": "艺术品",
        "exhibitions": "展览",
        "news": "新闻",
    }
    for key in ("works", "exhibitions", "news"):
        ws = wb.create_sheet(sheet_names[key])
        ws.append(headers)
        for item in copy_log[key]:
            ws.append([
                item["id"],
                item["title"],
                item["folder"],
                item["copied_count"],
                item["supplement_count"],
            ])

    notes = wb.create_sheet("说明")
    notes.append(["说明"])
    notes.append(["本索引用于帮助客户按编号和中文标题查阅资料。"])
    notes.append(["所有文件均为复制件，原网站仓库内容未被修改。"])
    notes.append(["如需查看未归档的补充原始资料，请打开 90_补充原始资料_网站全部资料原样备份。"])

    out_path = dest_root / "09_查阅索引" / "张满堂资料索引_客户版.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_root_readme(dest_root: Path) -> None:
    text = """张满堂资料总库（客户版）使用说明

1. 01_网站展示版
   当前网站展示所需的主要复制件，可用于整体浏览网站结构与内容。

2. 02_首页与品牌资料、03_作者介绍资料、07_联系资料
   存放首页、作者介绍、联系信息等站点基础资料。

3. 04_艺术品资料、05_展览资料、06_新闻动态资料
   每个条目均按“编号 + 中文标题”命名。
   文件夹内包含：
   - 资料卡
   - 网站中已整理好的图片或视频复制件
   - 能对应上的补充原始资料

4. 08_项目文档与设计系统
   存放网站文档、设计系统、内容表格等项目资料。

5. 09_查阅索引
   提供 Excel 索引，方便按编号和标题快速查找。

6. 90_补充原始资料_网站全部资料原样备份
   存放未结构化的完整补充原始资料备份，便于后续继续整理。
"""
    (dest_root / "00_请先阅读" / "客户版资料库使用说明.txt").parent.mkdir(parents=True, exist_ok=True)
    (dest_root / "00_请先阅读" / "客户版资料库使用说明.txt").write_text(text, encoding="utf-8")


def main() -> None:
    dest_root = ensure_unique_dir(DESTINATION_ROOT)
    dest_root.mkdir(parents=True, exist_ok=True)

    build_site_runtime(dest_root)
    build_profile_sections(dest_root)

    records = load_items()
    category_dirs = {
        "works": dest_root / "04_艺术品资料",
        "exhibitions": dest_root / "05_展览资料",
        "news": dest_root / "06_新闻动态资料",
    }
    for path in category_dirs.values():
        path.mkdir(parents=True, exist_ok=True)

    copy_log: dict[str, list[dict]] = defaultdict(list)
    for category, items in records.items():
        for record in items:
            folder, copied_names, supplement_names = copy_item_assets(record, category_dirs[category])
            copy_log[category].append(
                {
                    "id": record.item_id,
                    "title": record.title_zh,
                    "folder": str(folder),
                    "copied_count": len(copied_names),
                    "supplement_count": len(supplement_names),
                }
            )

    # Special unmatched supplemental news item
    qianlong_src = SUPPLEMENT_ROOT / "04_新闻动态_Credentials" / "北京千龙网"
    if qianlong_src.exists():
        copy_tree(qianlong_src, dest_root / "06_新闻动态资料" / "补充_北京千龙网视频")

    # Preserve complete raw artwork and supplemental backup for comprehensive review
    artwork_raw = SUPPLEMENT_ROOT / "03_艺术品_Artwork"
    if artwork_raw.exists():
        copy_tree(artwork_raw, dest_root / "04_艺术品资料" / "99_补充原始素材_来自网站全部资料")

    copy_full_supplement_backup(dest_root)
    write_root_readme(dest_root)
    create_index_workbook(dest_root, records, copy_log)

    print(str(dest_root).encode("unicode_escape").decode("ascii"))


if __name__ == "__main__":
    main()
