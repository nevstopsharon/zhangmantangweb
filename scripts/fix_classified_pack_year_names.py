from __future__ import annotations

from pathlib import Path
import re

from openpyxl import load_workbook


PACK_ROOT = Path(r"C:\Users\weixi\Desktop\张满堂资料总库_分类版")
SOURCE_WORKBOOK = Path(r"C:\Users\weixi\Desktop\zhangmantang-website\excel\content.xlsx")
INVALID_CHARS = re.compile(r'[<>:"/\\|?*]+')


def load_records(sheet_name: str, year_key: str):
    wb = load_workbook(SOURCE_WORKBOOK, data_only=True)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item_id = row[idx["id"]]
        if not item_id:
            continue
        rows.append(
            {
                "source_id": str(item_id),
                "year": str(row[idx[year_key]] or "").strip(),
                "title": str(row[idx["title_zh"]] or "").strip(),
            }
        )
    return rows


def safe_name(value: str) -> str:
    cleaned = INVALID_CHARS.sub(" ", str(value or "").strip())
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.strip().rstrip(".")


def sorted_item_dirs(base: Path, label: str) -> list[Path]:
    def key_fn(path: Path):
        m = re.match(rf"^{label}(\d+)", path.name)
        return int(m.group(1)) if m else 999999

    items = [
        p for p in base.iterdir()
        if p.is_dir() and not p.name.startswith("补充_") and not p.name.startswith("99_")
    ]
    return sorted(items, key=key_fn)


def rename_contents(folder: Path, label: str, num: str, year: str, title: str) -> None:
    new_prefix = f"{label}{num}_{year}_{safe_name(title)}"
    for item in list(folder.iterdir()):
        if not item.is_file():
            continue
        old_name = item.name
        if old_name.endswith("_资料卡.txt"):
            target = folder / f"{new_prefix}_资料卡.txt"
        elif "_封面." in old_name:
            target = folder / f"{new_prefix}_封面{item.suffix.lower()}"
        elif "_图片" in old_name:
            match = re.search(r"_图片(\d+)", old_name)
            idx = match.group(1) if match else "01"
            target = folder / f"{new_prefix}_图片{idx}{item.suffix.lower()}"
        else:
            match = re.match(r"^(work-|exhibition-|news-|作品|展览|新闻)\d+_(.+)$", old_name)
            if match:
                tail = match.group(2)
                target = folder / f"{new_prefix}_{tail}"
            else:
                continue
        if target != item:
            item.rename(target)


def rename_category(base: Path, label: str, records: list[dict], sequential: bool) -> None:
    folders = sorted_item_dirs(base, label)
    if len(folders) != len(records):
        raise RuntimeError(f"{label} 目录数量与记录数量不一致：{len(folders)} != {len(records)}")
    for idx, (folder, record) in enumerate(zip(folders, records), start=1):
        display_num = f"{idx:03d}" if sequential else record["source_id"].split("-")[-1]
        target_name = f"{label}{display_num}_{record['year']}_{safe_name(record['title'])}"
        target = folder.with_name(target_name)
        if target != folder:
            folder.rename(target)
            folder = target
        rename_contents(folder, label, display_num, record["year"], record["title"])


def main() -> None:
    exhibitions = load_records("exhibitions", "year")
    news = load_records("news", "date")
    rename_category(PACK_ROOT / "04_展览资料", "展览", exhibitions, sequential=False)
    rename_category(PACK_ROOT / "05_新闻动态资料", "新闻", news, sequential=True)
    print(str(PACK_ROOT).encode("unicode_escape").decode("ascii"))


if __name__ == "__main__":
    main()
