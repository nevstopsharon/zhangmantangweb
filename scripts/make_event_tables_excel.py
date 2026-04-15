from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


OUTPUT = Path(r"C:\Users\weixi\Desktop\商业文件\滨州泰安商会揭牌仪式表格.xlsx")


def apply_table_style(ws, max_row: int, max_col: int):
    thin = Side(style="thin", color="000000")
    header_fill = PatternFill("solid", fgColor="EDEDED")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.font = Font(name="Microsoft YaHei", size=11)
            cell.alignment = center

    for cell in ws[1]:
        cell.font = Font(name="Microsoft YaHei", size=11, bold=True)
        cell.fill = header_fill

    if max_col >= 3:
        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=max_col):
            for cell in row:
                cell.alignment = left


def build_workbook():
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "活动物料"
    ws1.append(["类别", "数量"])
    material_rows = [
        ["道旗", ""],
        ["拱门", "1个"],
        ["红毯", "20米"],
        ["签到合影墙", "1面"],
        ["指引牌", "2个"],
        ["签到桌", "2张"],
        ["影音设备", "1套"],
        ["条幅", "20米"],
        ["舞台挡板", ""],
        ["讲话台装饰", "1套"],
        ["启动道具", "1套"],
        ["奖杯奖牌", ""],
        ["定制礼品", "100套"],
    ]
    for row in material_rows:
        ws1.append(row)
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 14
    apply_table_style(ws1, ws1.max_row, ws1.max_column)

    wb.save(OUTPUT)


if __name__ == "__main__":
    build_workbook()
    print("generated")
