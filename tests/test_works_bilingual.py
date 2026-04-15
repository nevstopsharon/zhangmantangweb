from __future__ import annotations

import json
import sys
import unittest
import uuid
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
TMP_ROOT = ROOT / "tmpshots" / "test-temp"
sys.path.insert(0, str(ROOT))

from scripts.excel_to_json import export_workbook  # noqa: E402


EXPECTED_HEADERS = [
    "id",
    "title_zh",
    "title_en",
    "year",
    "material_zh",
    "material_en",
    "project_zh",
    "project_en",
    "size_zh",
    "size_en",
    "location_zh",
    "location_en",
    "description_zh",
    "description_en",
    "inspiration_zh",
    "inspiration_en",
    "cover_image",
    "gallery_images",
]

EXPECTED_MATERIALS = {"紫砂", "瓷", "宣纸", "其他"}
EXPECTED_PROJECTS = {
    "欧亚国家名片-中国文化邮票",
    "走向胜利",
    "人民大会堂藏品",
    "毛泽东诗词精选",
}


class WorksBilingualTests(unittest.TestCase):
    def make_temp_dir(self) -> Path:
        path = TMP_ROOT / str(uuid.uuid4())
        path.mkdir(parents=True, exist_ok=False)
        self.addCleanup(self.cleanup_dir, path)
        return path

    def make_project_paths(self) -> tuple[Path, Path, Path]:
        root = self.make_temp_dir()
        workbook_dir = root / "excel"
        output_dir = root / "out"
        workbook_dir.mkdir(parents=True, exist_ok=True)
        output_dir.mkdir(parents=True, exist_ok=True)
        return root, workbook_dir / "content.xlsx", output_dir

    def create_asset(self, root: Path, path: str) -> None:
        asset_path = root / path.lstrip("/")
        asset_path.parent.mkdir(parents=True, exist_ok=True)
        asset_path.write_bytes(b"test")

    def cleanup_dir(self, path: Path) -> None:
        if not path.exists():
            return
        for child in sorted(path.rglob("*"), reverse=True):
            if child.is_file():
                child.unlink()
            else:
                child.rmdir()
        path.rmdir()

    def test_workbook_uses_bilingual_works_headers(self) -> None:
        wb = load_workbook(ROOT / "excel" / "content.xlsx", data_only=True)
        ws = wb["works"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(headers, EXPECTED_HEADERS)

    def test_export_supports_bilingual_works_headers(self) -> None:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        works = workbook.create_sheet("works")
        works.append(EXPECTED_HEADERS)
        works.append(
            [
                "work-999",
                "测试作品",
                "Test Work",
                "2026",
                "宣纸",
                "Xuan Paper",
                "走向胜利",
                "Marching Toward Victory",
                "68x68厘米",
                "68 × 68 cm",
                "北京",
                "Beijing",
                "中文描述",
                "English description",
                "中文题记",
                "English inscription",
                "/images/works/work-999/cover.webp",
                "/images/works/work-999/detail-01.webp;/images/works/work-999/detail-02.webp",
            ]
        )

        exhibitions = workbook.create_sheet("exhibitions")
        exhibitions.append(
            [
                "id",
                "title_zh",
                "title_en",
                "year",
                "location_zh",
                "location_en",
                "description_zh",
                "description_en",
                "cover_image",
                "gallery_images",
            ]
        )

        news = workbook.create_sheet("news")
        news.append(
            [
                "id",
                "title_zh",
                "title_en",
                "date",
                "content_zh",
                "content_en",
                "related_report_label_zh",
                "related_report_label_en",
                "related_report_url",
                "full_video_label_zh",
                "full_video_label_en",
                "full_video_url",
                "cover_image",
                "gallery_images",
            ]
        )

        root, workbook_path, output_dir = self.make_project_paths()
        self.create_asset(root, "/images/works/work-999/cover.webp")
        self.create_asset(root, "/images/works/work-999/detail-01.webp")
        self.create_asset(root, "/images/works/work-999/detail-02.webp")
        workbook.save(workbook_path)
        export_workbook(workbook_path, output_dir)

        exported = json.loads((output_dir / "works.json").read_text(encoding="utf-8"))
        self.assertEqual(exported[0]["material_zh"], "宣纸")
        self.assertEqual(exported[0]["material_en"], "Xuan Paper")
        self.assertEqual(exported[0]["project_zh"], "走向胜利")
        self.assertEqual(exported[0]["project_en"], "Marching Toward Victory")
        self.assertEqual(exported[0]["size_zh"], "68x68厘米")
        self.assertEqual(exported[0]["size_en"], "68 × 68 cm")
        self.assertEqual(exported[0]["location_zh"], "北京")
        self.assertEqual(exported[0]["location_en"], "Beijing")
        self.assertEqual(
            exported[0]["gallery_images"],
            [
                "/images/works/work-999/detail-01.webp",
                "/images/works/work-999/detail-02.webp",
            ],
        )

    def test_exported_works_cover_required_filter_dimensions_and_paths(self) -> None:
        tmpdir = self.make_temp_dir()
        output_dir = tmpdir / "out"
        export_workbook(ROOT / "excel" / "content.xlsx", output_dir)
        works = json.loads((output_dir / "works.json").read_text(encoding="utf-8"))

        materials = {item["material_zh"] for item in works if item.get("material_zh")}
        projects = {item["project_zh"] for item in works if item.get("project_zh")}

        self.assertTrue(EXPECTED_MATERIALS.issubset(materials))
        self.assertTrue(EXPECTED_PROJECTS.issubset(projects))

        for item in works:
            self.assertTrue(item.get("title_en"), f"{item['id']} missing title_en")
            if item.get("description_zh"):
                self.assertTrue(item.get("description_en"), f"{item['id']} missing description_en")
            if item.get("inspiration_zh"):
                self.assertTrue(item.get("inspiration_en"), f"{item['id']} missing inspiration_en")

            for key in ("material_en", "project_en", "size_en", "location_en"):
                zh_key = key.replace("_en", "_zh")
                if item.get(zh_key):
                    self.assertTrue(item.get(key), f"{item['id']} missing {key}")

            for path in [item.get("cover_image"), *(item.get("gallery_images") or [])]:
                if not path:
                    continue
                self.assertTrue((ROOT / Path(path.lstrip("/"))).exists(), f"Missing asset: {path}")

    def test_export_normalizes_work_year_values(self) -> None:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        works = workbook.create_sheet("works")
        works.append(EXPECTED_HEADERS)
        works.append(
            [
                "work-1000",
                "年份测试",
                "Year Test",
                "2024年",
                "瓷",
                "Porcelain",
                "人民大会堂藏品",
                "Great Hall of the People Collection",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "/images/works/work-1000/cover.webp",
                "",
            ]
        )

        exhibitions = workbook.create_sheet("exhibitions")
        exhibitions.append(
            [
                "id",
                "title_zh",
                "title_en",
                "year",
                "location_zh",
                "location_en",
                "description_zh",
                "description_en",
                "cover_image",
                "gallery_images",
            ]
        )

        news = workbook.create_sheet("news")
        news.append(
            [
                "id",
                "title_zh",
                "title_en",
                "date",
                "content_zh",
                "content_en",
                "related_report_label_zh",
                "related_report_label_en",
                "related_report_url",
                "full_video_label_zh",
                "full_video_label_en",
                "full_video_url",
                "cover_image",
                "gallery_images",
            ]
        )

        root, workbook_path, output_dir = self.make_project_paths()
        self.create_asset(root, "/images/works/work-1000/cover.webp")
        workbook.save(workbook_path)
        export_workbook(workbook_path, output_dir)

        exported = json.loads((output_dir / "works.json").read_text(encoding="utf-8"))
        self.assertEqual(exported[0]["year"], "2024")


if __name__ == "__main__":
    unittest.main()
