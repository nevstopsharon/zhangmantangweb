from __future__ import annotations

import json
import sys
import unittest
import uuid
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
TMP_ROOT = ROOT / "tmpshots" / "test-temp"
sys.path.insert(0, str(ROOT))

from scripts.excel_to_json import export_workbook  # noqa: E402


EXPECTED_HEADERS = [
    "id",
    "title_zh",
    "title_en",
    "date",
    "content_zh",
    "content_en",
    "related_report_label_zh",
    "related_report_label_en",
    "related_report_url",
    "full_video_label_zh",
    "full_video_label_en",
    "full_video_url",
    "cover_image",
    "gallery_images",
]


class NewsOptionalLinksTests(unittest.TestCase):
    def make_temp_dir(self) -> Path:
        path = TMP_ROOT / str(uuid.uuid4())
        path.mkdir(parents=True, exist_ok=False)
        self.addCleanup(self.cleanup_dir, path)
        return path

    def make_project_paths(self) -> tuple[Path, Path, Path]:
        root = self.make_temp_dir()
        workbook_dir = root / "excel"
        output_dir = root / "out"
        workbook_dir.mkdir(parents=True, exist_ok=True)
        output_dir.mkdir(parents=True, exist_ok=True)
        return root, workbook_dir / "content.xlsx", output_dir

    def create_asset(self, root: Path, path: str) -> None:
        asset_path = root / path.lstrip("/")
        asset_path.parent.mkdir(parents=True, exist_ok=True)
        asset_path.write_bytes(b"test")

    def cleanup_dir(self, path: Path) -> None:
        if not path.exists():
            return
        for child in sorted(path.rglob("*"), reverse=True):
            if child.is_file():
                child.unlink()
            else:
                child.rmdir()
        path.rmdir()

    def test_workbook_uses_news_optional_link_headers(self) -> None:
        wb = load_workbook(ROOT / "excel" / "content.xlsx", data_only=True)
        ws = wb["news"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertEqual(headers, EXPECTED_HEADERS)

    def test_export_supports_news_optional_link_fields(self) -> None:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        works = workbook.create_sheet("works")
        works.append(
            [
                "id",
                "title_zh",
                "title_en",
                "year",
                "material_zh",
                "material_en",
                "project_zh",
                "project_en",
                "size_zh",
                "size_en",
                "location_zh",
                "location_en",
                "description_zh",
                "description_en",
                "inspiration_zh",
                "inspiration_en",
                "cover_image",
                "gallery_images",
            ]
        )

        exhibitions = workbook.create_sheet("exhibitions")
        exhibitions.append(
            [
                "id",
                "title_zh",
                "title_en",
                "year",
                "location_zh",
                "location_en",
                "description_zh",
                "description_en",
                "cover_image",
                "gallery_images",
            ]
        )

        news = workbook.create_sheet("news")
        news.append(EXPECTED_HEADERS)
        news.append(
            [
                "news-999",
                "测试新闻",
                "Test News",
                "2026-01-01",
                "中文正文",
                "English body",
                "相关报道：",
                "Related coverage:",
                "https://example.com/report",
                "完整视频在：",
                "Full video:",
                "https://example.com/video",
                "/images/news/news-999/cover.webp",
                "/images/news/news-999/detail-01.webp;/images/news/news-999/detail-02.webp",
            ]
        )

        root, workbook_path, output_dir = self.make_project_paths()
        self.create_asset(root, "/images/news/news-999/cover.webp")
        self.create_asset(root, "/images/news/news-999/detail-01.webp")
        self.create_asset(root, "/images/news/news-999/detail-02.webp")
        workbook.save(workbook_path)
        export_workbook(workbook_path, output_dir)

        exported = json.loads((output_dir / "news.json").read_text(encoding="utf-8"))
        self.assertEqual(exported[0]["related_report_label_zh"], "相关报道：")
        self.assertEqual(exported[0]["related_report_label_en"], "Related coverage:")
        self.assertEqual(exported[0]["related_report_url"], "https://example.com/report")
        self.assertEqual(exported[0]["full_video_label_zh"], "完整视频在：")
        self.assertEqual(exported[0]["full_video_label_en"], "Full video:")
        self.assertEqual(exported[0]["full_video_url"], "https://example.com/video")
        self.assertEqual(
            exported[0]["gallery_images"],
            [
                "/images/news/news-999/detail-01.webp",
                "/images/news/news-999/detail-02.webp",
            ],
        )

    def test_export_normalizes_news_dates(self) -> None:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        works = workbook.create_sheet("works")
        works.append(
            [
                "id",
                "title_zh",
                "title_en",
                "year",
                "material_zh",
                "material_en",
                "project_zh",
                "project_en",
                "size_zh",
                "size_en",
                "location_zh",
                "location_en",
                "description_zh",
                "description_en",
                "inspiration_zh",
                "inspiration_en",
                "cover_image",
                "gallery_images",
            ]
        )

        exhibitions = workbook.create_sheet("exhibitions")
        exhibitions.append(
            [
                "id",
                "title_zh",
                "title_en",
                "year",
                "location_zh",
                "location_en",
                "description_zh",
                "description_en",
                "cover_image",
                "gallery_images",
            ]
        )

        news = workbook.create_sheet("news")
        news.append(EXPECTED_HEADERS)
        news.append(
            [
                "news-1000",
                "日期测试",
                "Date Test",
                "2024年3月8日",
                "中文正文",
                "English body",
                "",
                "",
                "",
                "",
                "",
                "",
                "/images/news/news-1000/cover.webp",
                "",
            ]
        )

        root, workbook_path, output_dir = self.make_project_paths()
        self.create_asset(root, "/images/news/news-1000/cover.webp")
        workbook.save(workbook_path)
        export_workbook(workbook_path, output_dir)

        exported = json.loads((output_dir / "news.json").read_text(encoding="utf-8"))
        self.assertEqual(exported[0]["date"], "2024-03-08")


if __name__ == "__main__":
    unittest.main()
