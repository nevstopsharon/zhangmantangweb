from __future__ import annotations

import sys
import unittest
import uuid
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
TMP_ROOT = ROOT / "tmpshots" / "test-temp"
sys.path.insert(0, str(ROOT))

from scripts.excel_to_json import export_workbook  # noqa: E402


WORK_HEADERS = [
    "id",
    "title_zh",
    "title_en",
    "year",
    "material_zh",
    "material_en",
    "project_zh",
    "project_en",
    "size_zh",
    "size_en",
    "location_zh",
    "location_en",
    "description_zh",
    "description_en",
    "inspiration_zh",
    "inspiration_en",
    "cover_image",
    "gallery_images",
]

EXHIBITION_HEADERS = [
    "id",
    "title_zh",
    "title_en",
    "year",
    "location_zh",
    "location_en",
    "description_zh",
    "description_en",
    "cover_image",
    "gallery_images",
]

NEWS_HEADERS = [
    "id",
    "title_zh",
    "title_en",
    "date",
    "content_zh",
    "content_en",
    "related_report_label_zh",
    "related_report_label_en",
    "related_report_url",
    "full_video_label_zh",
    "full_video_label_en",
    "full_video_url",
    "cover_image",
    "gallery_images",
]


class ExportValidationTests(unittest.TestCase):
    def make_temp_dir(self) -> Path:
        path = TMP_ROOT / str(uuid.uuid4())
        path.mkdir(parents=True, exist_ok=False)
        self.addCleanup(self.cleanup_dir, path)
        return path

    def cleanup_dir(self, path: Path) -> None:
        if not path.exists():
            return
        for child in sorted(path.rglob("*"), reverse=True):
            if child.is_file():
                child.unlink()
            else:
                child.rmdir()
        path.rmdir()

    def make_project_paths(self) -> tuple[Path, Path, Path]:
        root = self.make_temp_dir()
        workbook_dir = root / "excel"
        output_dir = root / "out"
        workbook_dir.mkdir(parents=True, exist_ok=True)
        output_dir.mkdir(parents=True, exist_ok=True)
        return root, workbook_dir / "content.xlsx", output_dir

    def make_workbook(self) -> Workbook:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        works = workbook.create_sheet("works")
        works.append(WORK_HEADERS)

        exhibitions = workbook.create_sheet("exhibitions")
        exhibitions.append(EXHIBITION_HEADERS)

        news = workbook.create_sheet("news")
        news.append(NEWS_HEADERS)

        return workbook

    def test_export_rejects_duplicate_ids_within_sheet(self) -> None:
        workbook = self.make_workbook()
        works = workbook["works"]
        works.append(
            [
                "work-dup",
                "作品一",
                "Work One",
                "2024",
                "瓷",
                "Porcelain",
                "人民大会堂藏品",
                "Great Hall of the People Collection",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )
        works.append(
            [
                "work-dup",
                "作品二",
                "Work Two",
                "2025",
                "瓷",
                "Porcelain",
                "人民大会堂藏品",
                "Great Hall of the People Collection",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ]
        )

        _, workbook_path, output_dir = self.make_project_paths()
        workbook.save(workbook_path)

        with self.assertRaisesRegex(ValueError, "Duplicate id"):
            export_workbook(workbook_path, output_dir)

    def test_export_rejects_missing_local_images(self) -> None:
        workbook = self.make_workbook()
        news = workbook["news"]
        news.append(
            [
                "news-missing-image",
                "图片测试",
                "Image Test",
                "2024-03-08",
                "中文正文",
                "English body",
                "",
                "",
                "",
                "",
                "",
                "",
                "/images/news/news-missing-image/cover.webp",
                "",
            ]
        )

        _, workbook_path, output_dir = self.make_project_paths()
        workbook.save(workbook_path)

        with self.assertRaisesRegex(ValueError, "Missing asset"):
            export_workbook(workbook_path, output_dir)
