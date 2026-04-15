from __future__ import annotations

import argparse
import json
import os
import re
from pathlib import Path

from openpyxl import load_workbook


SHEET_SPECS = {
    "works": [
        "id",
        "title_zh",
        "title_en",
        "year",
        "material_zh",
        "material_en",
        "project_zh",
        "project_en",
        "size_zh",
        "size_en",
        "location_zh",
        "location_en",
        "description_zh",
        "description_en",
        "inspiration_zh",
        "inspiration_en",
        "cover_image",
        "gallery_images",
    ],
    "exhibitions": [
        "id",
        "title_zh",
        "title_en",
        "year",
        "location_zh",
        "location_en",
        "description_zh",
        "description_en",
        "cover_image",
        "gallery_images",
    ],
    "news": [
        "id",
        "title_zh",
        "title_en",
        "date",
        "content_zh",
        "content_en",
        "related_report_label_zh",
        "related_report_label_en",
        "related_report_url",
        "full_video_label_zh",
        "full_video_label_en",
        "full_video_url",
        "cover_image",
        "gallery_images",
    ],
}


def split_multi_value(value: str | None) -> list[str]:
    if not value:
        return []
    parts = re.split(r"[;\n,]+", value)
    return [part.strip() for part in parts if part.strip()]


def normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def sheet_to_records(workbook_path: Path, sheet_name: str) -> list[dict]:
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Missing required sheet: {sheet_name}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_cell(value) for value in rows[0]]
    expected = SHEET_SPECS[sheet_name]
    if headers[: len(expected)] != expected:
        raise ValueError(f"Sheet '{sheet_name}' headers do not match expected fields.")

    records = []
    for raw_row in rows[1:]:
        if not raw_row or not any(cell is not None and str(cell).strip() for cell in raw_row):
            continue
        record = {}
        for index, key in enumerate(expected):
            record[key] = normalize_cell(raw_row[index] if index < len(raw_row) else "")
        record["gallery_images"] = split_multi_value(record.get("gallery_images"))
        records.append(record)
    return records


def export_workbook(workbook_path: Path, output_dir: Path) -> None:
    os.makedirs(output_dir, exist_ok=True)
    for sheet_name in SHEET_SPECS:
        records = sheet_to_records(workbook_path, sheet_name)
        (output_dir / f"{sheet_name}.json").write_text(
            json.dumps(records, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert CMS Excel sheets into JSON data files.")
    parser.add_argument("--workbook", type=Path, required=True, help="Path to excel/content.xlsx")
    parser.add_argument("--output-dir", type=Path, required=True, help="Output directory for JSON files.")
    args = parser.parse_args()

    export_workbook(args.workbook.resolve(), args.output_dir.resolve())
    print(f"Exported JSON to: {args.output_dir.resolve()}")


if __name__ == "__main__":
    main()
